import time
from constants import *
from classifiers import conduct_experiment, graph_experiment, plot_cutoff, Experiment
import urllib
import csv
from datetime import datetime, timedelta
import os
from buyers import check_prediction, check_prediction_ex
from utils import save, load, abs_path
import settings
from openpyxl import Workbook
import features
import numpy as np
from openpyxl.drawing.image import Image


db = settings.get_db()
collection = db.IntegratedDocument

userTypeId = "59cbc103003e730508e87c2c"
appId = "123123123"#
company = "Netinfo"
features_wrapper = features.FeaturesWrapper(appId)
case_filter ={
    "UserId": appId, "TypeId": userTypeId
}
weeksAvailable = collection.find(case_filter).distinct("Document.g_timestamp")
weeksAvailable.sort() 
print weeksAvailable
current_week = weeksAvailable[7]
weeks_count = 4
exp = Experiment(None, None, None, company)
models = exp.load_model_files(['rf'])
model_cutoffs = {'gba': 0.1, 'rf': 0.2 }
print "Loaded prediction models"
    # models = [
    #   Experiment.load_model("/experiments/Netinfo/model_gba_2017-10-04_12:44:31.pickle",company),
    #   Experiment.load_model("/experiments/Netinfo/model_rf_2017-10-04_14:57:34.pickle", company)
    #]
for week_ix in xrange(weeks_count):
    target_week = current_week
    target_week_end = target_week + timedelta(days=7)

    print "Target week {0} will predict for {1}".format(target_week, target_week_end)
    paying_users = collection.find(dict(case_filter, **{ "Document.is_paying": 1 })).distinct("Document.uuid")

    next_week = target_week_end
    next_week_end = target_week_end + timedelta(days=7)
    next_week_f = str(next_week).replace(':', '_')
    filter_entries = True 
    userFeatures = []
    userData = []
    targets = []

    print "Gathering data from week " + str(target_week)
    prediction_features_query = dict(case_filter, **{"Document.noticed_date": {'$gte': target_week, '$lt': target_week_end}})
    pipeline = [
            {"$match": prediction_features_query},
            {"$group": {
                "_id": { "uuid" : "$Document.uuid"},
                "visits_count": {"$sum": 1}
            }
            }]
    pipeline[1]["$group"].update(features_wrapper.get_feature_avg_aggregates("$Document.")) 
    weekData = collection.aggregate(pipeline, allowDiskUse=True)
    weekData = list(weekData) 
    previously_purchasing_users = collection.find(dict(case_filter, **{
        "Document.noticed_date": {"$lt": target_week_end}, "Document.is_paying": 1
    })).distinct("Document.uuid")

    next_week_purchasing_users = users = collection.find(dict(case_filter, **{ 
        "Document.is_paying": 1, "Document.noticed_date": {'$gte': next_week, '$lte': next_week_end}
    })).distinct("Document.uuid")

    for tmpDoc in weekData:
        uuid = tmpDoc["_id"]["uuid"]
        if uuid in previously_purchasing_users: continue # skip users that have paid some time
        userData.append({'uuid': uuid})
        # since we have only 1 week
        tmpDoc["has_paid_before"] = 1 if uuid in previously_purchasing_users else 0 
        userFeatures.append(features_wrapper.get_values(tmpDoc))
        target_val = 0
        if "Document.is_paying" in prediction_features_query and prediction_features_query["Document.is_paying"] == 1:
            target_val = 1
        else:
            if uuid in next_week_purchasing_users and tmpDoc["has_paid_before"] == 0:
                target_val = 1
        targets.append(target_val)

    print "Loaded " + str(len(userFeatures)) + " test user items" 
    payingDict = dict() 

    filtered = 0
    for m in models:
        t_started = time.time()
        c_type = m['type']    # x_train = Experiment.load_dump(company, 'x_train_{0}.dmp'.format(c_type))  y_train = Experiment.load_dump(company, 'y_train_{0}.dmp'.format(c_type)) 
        predictions = m['model'].predict_proba(userFeatures)
        x_test = Experiment.load_dump(company, 'x_test_{0}.dmp'.format(c_type))
        y_true = Experiment.load_dump(company, 'y_true_{0}.dmp'.format(c_type))
        real_cutoff = plot_cutoff(m, x_test, y_true, client=company)
        prediction_graphics = graph_experiment(userFeatures, targets, company, m)
        prediction_graphics[c_type]['importance'] = ex.predict_explain_non_tree2(m, company, features_wrapper.get_names())
        
        fileName = os.path.join(company, 'prediction_{0}_{1}.xlsx'.format(c_type, next_week_f))
        print "Writing predictions in: " + fileName
        report = Workbook()
        main_sheet = report.active
        info_sheet = report.create_sheet("Output")
        validations_sheet = report.create_sheet("Validated")
        graphics = report.create_sheet('graphics')
        graphics.append([ "Graphics for {0}".format(c_type) ])
        graphics.add_image(Image(prediction_graphics[c_type]['importance']), 'A1')
        graphics.add_image(Image(prediction_graphics[c_type]['confusion']), 'A5')

        main_sheet.append(["uuid", "chance to by"])
        for p in xrange(len(predictions)):
            user_pred = predictions[p]
            uuid = userData[p]['uuid']
            main_sheet.append([uuid, user_pred[1]])
        time_taken = time.time() - t_started
        info_sheet.append([
            "{0} Duration of prediction: {1} ({2:0.4f}s/u)".format(c_type, time_taken, time_taken / len(predictions))
        ]) 
        cutoff_value = model_cutoffs[c_type]
        check = check_prediction_ex(next_week, {
            'users' : userData, 'predictions': predictions
        }, cutoff_value, company, c_type)
        for match in check['matches']:
            uuid = match['uuid']
            chance = float(match['chance'])
            validations_sheet.append([uuid, chance, chance >= cutoff_value])

        #rf Matches 90.1961%: 51, pos46 c0.2(0.6974)rf False positives: 14.0023%, count 3084 of 22025 with (0.0685-0.3187)% chance
        # print "Users paid in {0}, but not in the previous week: {1} ({2} filtered)".format(next_week, len(check['payers']), check['filtered'])
        str_matches = "{0:0.4f}% ({1} of {2})".format(check['positive_perc'], check['positive_matches'], len(check['matches']))
        str_cutoff = "{0:0.4f}(highest {1:0.4f})".format(cutoff_value, real_cutoff)
        str_fp = "{0:0.4f}% ({1} of {2} / med {3:0.4f} - {4:0.4f})".format(
            check['false_positives']['perc'],
            check['false_positives']['count'],
            check['false_positives']['out_of'],
            check['false_positives']['chance']['median'],
            check['false_positives']['chance']['high'])
        info_sheet.append(["Matches",str_matches])
        info_sheet.append(["Cutoff",str_cutoff])
        info_sheet.append(["FP", str_fp])
        info_sheet.append(["Precision", "{0:0.4f}".format(check['precision'])])
        info_sheet.append(["Recall", "{0:0.4f}".format(check['recall'])])
        print "Saving output in {0}".format(fileName)
        report.save(fileName)
        print "Matches {0}\nCutoff {1}\n FP {2}\n".format(str_matches, str_cutoff, str_fp)
        nowVal = datetime.now().strftime("%Y-%m-%d_%H_%M_%S")
      
    current_week = target_week_end
 